from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
from catalog.models import Product
from account.models import User
from account.decorators import login_required_custom
from .models import Receipt, ReceiptItem, Supply, SupplyItem
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
import uuid


def gen_number(prefix):
    return f'{prefix}-{timezone.now().strftime("%Y%m%d")}-{str(uuid.uuid4())[:4].upper()}'


@login_required_custom
def receipt_list(request):
    receipts = Receipt.objects.all()
    sort = request.GET.get('sort', '-created_at')
    if sort == 'date_asc':
        receipts = receipts.order_by('created_at')
    elif sort == 'date_desc':
        receipts = receipts.order_by('-created_at')
    elif sort == 'total_asc':
        receipts = receipts.order_by('total')
    elif sort == 'total_desc':
        receipts = receipts.order_by('-total')
    else:
        receipts = receipts.order_by('-created_at')
    return render(request, 'receipt/receipt_list.html', {'receipts': receipts, 'sort': sort})


@login_required_custom
def receipt_create(request):
    cart = request.session.get('cart', {})
    items = []
    total = 0
    for pid, qty in cart.items():
        try:
            p = Product.objects.get(id=int(pid))
            subtotal = p.final_price * qty
            total += subtotal
            items.append({'product': p, 'qty': qty, 'subtotal': subtotal, 'price': p.final_price})
        except Product.DoesNotExist:
            pass

    if request.method == 'POST':
        user_id = request.session.get('user_id')
        seller = User.objects.get(id=user_id) if user_id else None
        receipt = Receipt.objects.create(
            number=gen_number('CHK'),
            seller=seller,
            status='draft',
            total=total,
        )
        for item in items:
            ReceiptItem.objects.create(
                receipt=receipt,
                product=item['product'],
                product_name=item['product'].name,
                price=item['price'],
                quantity=item['qty'],
                subtotal=item['subtotal'],
            )
        receipt.calc_total()
        request.session['cart'] = {}
        request.session.modified = True
        messages.success(request, f'Чек {receipt.number} создан')
        return redirect('receipt:receipt_detail', pk=receipt.pk)
    return render(request, 'receipt/receipt_create.html', {'items': items, 'total': total})


@login_required_custom
def receipt_detail(request, pk):
    receipt = get_object_or_404(Receipt, pk=pk)
    return render(request, 'receipt/receipt_detail.html', {'receipt': receipt})


@login_required_custom
def receipt_confirm(request, pk):
    receipt = get_object_or_404(Receipt, pk=pk)
    if receipt.status == 'draft':
        for item in receipt.items.all():
            if item.product:
                item.product.stock -= item.quantity
                item.product.sold_count += item.quantity
                item.product.save()
        receipt.status = 'confirmed'
        receipt.save()
        messages.success(request, 'Чек проведён')
    return redirect('receipt:receipt_detail', pk=pk)


@login_required_custom
def receipt_export(request, pk):
    receipt = get_object_or_404(Receipt, pk=pk)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Чек {receipt.number}'

    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    center = Alignment(horizontal='center', vertical='center')
    accent_fill = PatternFill(start_color='7161EF', end_color='7161EF', fill_type='solid')
    white_font = Font(bold=True, color='FFFFFF', size=11)

    ws.merge_cells('A1:F1')
    ws['A1'] = 'АВТОЗАПЧАСТИ — ТОВАРНЫЙ ЧЕК'
    ws['A1'].font = title_font
    ws['A1'].alignment = center

    ws.merge_cells('A2:F2')
    ws['A2'] = f'Чек № {receipt.number}  |  Дата: {receipt.created_at.strftime("%d.%m.%Y %H:%M")}  |  Продавец: {receipt.seller.full_name if receipt.seller else "—"}'
    ws['A2'].alignment = center

    ws.row_dimensions[4].height = 20
    headers = ['№', 'Наименование', 'Кол-во', 'Цена', 'Сумма', 'Примечание']
    cols = ['A', 'B', 'C', 'D', 'E', 'F']
    for i, (col, h) in enumerate(zip(cols, headers)):
        cell = ws[f'{col}4']
        cell.value = h
        cell.font = white_font
        cell.fill = accent_fill
        cell.alignment = center
        cell.border = border

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    for idx, item in enumerate(receipt.items.all(), 1):
        row = idx + 4
        data = [idx, item.product_name, item.quantity, float(item.price), float(item.subtotal), '']
        for col, val in zip(cols, data):
            cell = ws[f'{col}{row}']
            cell.value = val
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    total_row = receipt.items.count() + 5
    ws.merge_cells(f'A{total_row}:D{total_row}')
    ws[f'A{total_row}'] = 'ИТОГО:'
    ws[f'A{total_row}'].font = Font(bold=True, size=12)
    ws[f'A{total_row}'].alignment = Alignment(horizontal='right')
    ws[f'E{total_row}'] = float(receipt.total)
    ws[f'E{total_row}'].font = Font(bold=True, size=12)
    ws[f'E{total_row}'].border = border

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="receipt_{receipt.number}.xlsx"'
    wb.save(response)
    return response


@login_required_custom
def supply_list(request):
    supplies = Supply.objects.all()
    sort = request.GET.get('sort', '-created_at')
    if sort == 'date_asc':
        supplies = supplies.order_by('created_at')
    elif sort == 'date_desc':
        supplies = supplies.order_by('-created_at')
    elif sort == 'total_asc':
        supplies = supplies.order_by('total')
    elif sort == 'total_desc':
        supplies = supplies.order_by('-total')
    else:
        supplies = supplies.order_by('-created_at')
    return render(request, 'receipt/supply_list.html', {'supplies': supplies, 'sort': sort})


@login_required_custom
def supply_detail(request, pk):
    supply = get_object_or_404(Supply, pk=pk)
    return render(request, 'receipt/supply_detail.html', {'supply': supply})
